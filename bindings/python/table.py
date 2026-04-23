"""Utilities for reading and writing DLite instances from and to tables."""
import csv
import re
from pathlib import Path
from typing import TYPE_CHECKING

import dlite

if TYPE_CHECKING:  # pragma: no cover
    from typing import Optional, Sequence


# Default mappings of DLite metadata fields to table header names
DEFAULT_DATAMODEL_MAPPINGS = {
    "uri": "@id",
    "dimensions": None,
    "description": "description",
}
# Default mappings of DLite property fields to table header names
DEFAULT_PROPERTY_MAPPINGS = {
    "name": "datumName",
    "type": "datumType",
    "ref": "datumRef",
    "unit": "datumUnit",
    "shape": "datumShape",
    "description": "datumDescription",
    "mapping": "datumMapping",
}


class DMTable():
    """A class for loading DLite data models form tables.

    A table with datamodels should contain some columns with pre-defined
    header labels. For details, see the [online documentation].

    References:
    [online documentation]: https://sintef.github.io/dlite/user_guide/datamodels.html#defining-multiple-datamodels-in-a-table
    """

    def __init__(
        self,
        table: "Sequence[Sequence]",
        datamodel_mappings: dict = DEFAULT_DATAMODEL_MAPPINGS,
        property_mappings: dict = DEFAULT_PROPERTY_MAPPINGS,
        baseuri: "Optional[str]" = None,
    ) -> None:
        """Initialises a DMTable object from a list of lists.

        Arguments:
            table: Table to load, represented as a sequence of sequences.
            datamodel_mappings: Mapping of DLite datamodel fields (uri,
                dimensions, description) to table header names. 'dimensions'
                is normally not provided, in which case it will be inferred
                from `property_mappings`.
            property_mappings: Mapping of DLite property fields (name, type,
                ref, unit, shape, description) to table header names.
            baseuri: Base URI to use if the data model identifier (the column
                that "uri" is mapped to) has no namespace.
                For example, if `baseuri="http://example.com/data/0.1/"` for
                a column with identifier "blah" will result in the URI
                "http://example.com/data/0.1/blah".

        """
        self.datamodels = {}  # Maps uri to datamodel dict
        self.datamodel_mappings = datamodel_mappings
        self.property_mappings = property_mappings

        # Get the indices of the columns that matches the mapped labels
        header = [h.strip() for h in table[0]]
        datamodel_idict = self._get_datamodel_idict(header)
        property_idicts = self._get_property_idicts(header)

        for row in table[1:]:
            d = {}

            # Parse datamodel mappings
            for k, i in datamodel_idict.items():
                d[k] = row[i].strip()
            if not re.match("^[a-z]+://", d["uri"]):
                if not baseuri:
                    raise ValueError(
                        f"Datamodel '{d['uri']}' has no namespace. "
                        "Provide a default namespace with the `baseuri` "
                        "argument."
                    )
                d["uri"] = baseuri + d["uri"]

            # Parse property mappings
            dims = {}
            for idict in property_idicts:
                prop = {}
                for k, i in idict.items():
                    value = row[i].strip() if row[i] else ""

                    if k == "shape" and value:
                            prop[k] = [
                                s.strip()
                                for s in value.strip("[]").split(",")
                            ]
                            for dim in prop[k]:
                                dims[dim] = f"{dim} dimension"
                    elif value:
                        prop[k] = value
                if prop.get("name"):
                    if "properties" in d:
                        d["properties"].append(prop)
                    else:
                        d["properties"] = [prop]
            if dims:
                d["dimensions"] = dims

            self.datamodels[d["uri"]] = d

    def _get_datamodel_idict(self, header: "Sequence[str]") -> "dict":
        """Help function that returns a dict mapping datamodel fields to
        corresponding header indices."""
        revmap = {v: k for k, v in self.datamodel_mappings.items()}
        d = {}
        for i, headname in enumerate(header):
            if headname in revmap:
                d[revmap[headname]] = i
        return d

    def _get_property_idicts(self, header: "Sequence[str]") -> "list[dict]":
        """Help function that returns a list of dicts mapping property fields to
        corresponding header indices."""
        typelabel = self.property_mappings["type"]
        headidx = {h: i for i, h in enumerate(header)}

        # Search the header for all names starting with the type label
        # (optionally followed by a bracket).
        # Not implemented now, but it is in principle possible to infer the
        # name label from the bracket.
        brackets = []
        for h in header:
            m = re.match(rf"^{typelabel}(\[[^]]*\])?$", h)
            if m:
                bracket, = m.groups()
                brackets.append("" if bracket is None else bracket)

        idicts = []
        for bracket in brackets:
            d = {}
            for name, headname in self.property_mappings.items():
                if headname+bracket in headidx:
                    d[name] = headidx[headname+bracket]
            idicts.append(d)

        return idicts

    def get_datamodels(self) -> "list[dlite.Metadata]":
        """Return a list with all datamodels parsed from the table."""
        return [dlite.Metadata.from_dict(d) for d in self.datamodels.values()]

    @staticmethod
    def from_csv(
        csvfile: "Union[Iterable[str], Path, str]",
        encoding: str = "utf-8",
        dialect: "Optional[Union[csv.Dialect, str]]" = None,
        datamodel_mappings: dict = DEFAULT_DATAMODEL_MAPPINGS,
        property_mappings: dict = DEFAULT_PROPERTY_MAPPINGS,
        baseuri: "Optional[str]" = None,
        **kwargs,
    ) -> "DMTable":
        # pylint: disable=line-too-long
        """Parse a csv file using the standard library csv module.

        Arguments:
            csvfile: Name of CSV file to parse or an iterable of strings.
            encoding: The encoding of the csv file.  Note that Excel may
                encode as "ISO-8859" (which was commonly used in the 1990s).
            dialect: A subclass of csv.Dialect, or the name of the dialect,
                specifying how the `csvfile` is formatted.  For more details,
                see [Dialects and Formatting Parameters].
            datamodel_mappings: Mapping of DLite datamodel fields (uri,
                dimensions, description) to table header names. 'dimensions'
                is normally not provided, in which case it will be inferred
                from `property_mappings`.
            property_mappings: Mapping of DLite property fields (name, type,
                ref, unit, shape, description) to table header names.
            baseuri: Base URI to use if the identifier has no namespace.
                See the corresponding argument of the __init__() method for
                details.
            kwargs: Additional keyword arguments overriding individual
                formatting parameters.  For more details, see
                [Dialects and Formatting Parameters].

        Returns:
            New DMTable instance.

        References:
        [Dialects and Formatting Parameters]: https://docs.python.org/3/library/csv.html#dialects-and-formatting-parameters
        """

        def read(f, dialect):
            """Return csv reader from file-like object `f`."""
            if dialect is None and not kwargs:
                sample = f.read(1024)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t ")
                except csv.Error:
                    # The build-in sniffer not always work well with
                    # non-numerical csv files. Try our simple sniffer
                    dialect = csvsniff(sample)
                finally:
                    f.seek(0)
            reader = csv.reader(f, dialect=dialect, **kwargs)
            return list(reader)

        if isinstance(csvfile, (str, Path)):
            with open(csvfile, mode="rt", encoding=encoding) as f:
                table = read(f, dialect)
        else:
            table = read(csvfile, dialect)

        return DMTable(
            table=table,
            datamodel_mappings=datamodel_mappings,
            property_mappings=property_mappings,
            baseuri=baseuri,
        )

    @staticmethod
    def from_excel(
        excelfile: "Union[Path, str]",
        sheet: "Union[str, int]" = 0,
        cellrange: "Optional[str]" = None,
        datamodel_mappings: dict = DEFAULT_DATAMODEL_MAPPINGS,
        property_mappings: dict = DEFAULT_PROPERTY_MAPPINGS,
        baseuri: "Optional[str]" = None,
        **kwargs,
    ) -> "DMTable":
        """Parse a csv file using the standard library csv module.

        Arguments:
            excelfile: Name of excel file to parse.
            sheet: Sheet name or number to load.
            cellrange: Cell range to load. Examples: "A1:C4", "A:C", "1:4".
                The default is to read all cells.
            dialect: A subclass of csv.Dialect, or the name of the dialect,
                specifying how the `csvfile` is formatted.  For more details,
                see [Dialects and Formatting Parameters].
            datamodel_mappings: Mapping of DLite datamodel fields (uri,
                dimensions, description) to table header names. 'dimensions'
                is normally not provided, in which case it will be inferred
                from `property_mappings`.
            property_mappings: Mapping of DLite property fields (name, type,
                ref, unit, shape, description) to table header names.
            baseuri: Base URI to use if the identifier has no namespace.
                See the corresponding argument of the __init__() method for
                details.
            kwargs: Additional keyword arguments overriding individual
                formatting parameters.  For more details, see
                [Dialects and Formatting Parameters].

        Returns:
            New DMTable instance.

        """
        from openpyxl import load_workbook

        wb = load_workbook(
            excelfile,
            read_only=True,
            keep_vba=False,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )
        # Get worksheet
        ws = wb[wb.sheetnames[sheet] if isinstance(sheet, int) else sheet]

        # Get cell range
        if cellrange:
            cr = ws[cellrange]
        else:
            # Find first non-empty rows and columns
            nrows = next((i for i, r in enumerate(ws.values) if not r[0]), 0)
            ncols = next((i for i, v in enumerate(next(ws.values)) if not v), 0)
            cr = ws.iter_rows(max_row=nrows, max_col=ncols)

        table = [[cell.value for cell in row] for row in cr]

        return DMTable(
            table = table,
            datamodel_mappings=datamodel_mappings,
            property_mappings=property_mappings,
            baseuri=baseuri,
        )

    def to_triplestore(self, ts: "Triplestore"):
        """Save all datamodels to Tripper triplestore `ts`."""

        # Import here since since dlite.dataset depends on tripper
        from dlite.dataset import add_dataset

        for uri, d in self.dmdicts.items():
            meta = dlite.Metadata.from_dict(d)
            mappings = []
            for props in d["properties"]:
                if "mapping" in props:
                    iri = f"{uri}#{props['name']}"
                    mappings.append((iri, "rdfs:subClassOf", props["mapping"]))
            add_dataset(ts, meta, iri=uri, mappings=mappings)


def csvsniff(sample: str) -> "csv.Dialect":
    """Custom csv sniffer.

    Analyse csv sample and returns a csv.Dialect instance.
    """
    # Determine line terminator
    if "\r\n" in sample:
        linesep = "\r\n"
    else:
        counts = {s: sample.count(s) for s in "\n\r"}
        linesep = max(counts, key=lambda k: counts[k])

    lines = sample.split(linesep)
    del lines[-1]  # skip last line since it might be truncated
    if not lines:
        raise csv.Error(
            "too long csv header. No line terminator within sample"
        )
    header = lines[0]

    # Possible delimiters and quote chars to check
    delims = [d for d in ",;\t :" if header.count(d)]
    quotes = [q for q in "\"'" if sample.count(q)]
    if not quotes:
        quotes = ['"']

    # For each (quote, delim)-pair, count the number of tokens per line
    # Only pairs for which all lines has the same number of tokens are added
    # to ntokens
    ntokens = {}  # map (quote, delim) to number of tokens per line
    for q in quotes:
        for d in delims:
            ntok = []
            for ln in lines:
                # Remove quoted tokens
                ln = re.sub(f"(^{q}[^{q}]*{q}{d})|({d}{q}[^{q}]*{q}$)", d, ln)
                ln = re.sub(f"{d}{q}[^{q}]*{q}{d}", d * 2, ln)
                ntok.append(len(ln.split(d)))

            if ntok and max(ntok) == min(ntok):
                ntokens[(q, d)] = ntok[0]

    # From ntokens, select (quote, delim) pair that results in the highest
    # number of tokens per line
    if not ntokens:
        raise csv.Error("not able to determine delimiter")
    quote, delim = max(ntokens, key=lambda k: ntokens[k])

    class dialect(csv.Dialect):
        """Custom dialect."""

        # pylint: disable=too-few-public-methods
        _name = "sniffed"
        delimiter = delim
        doublequote = True  # quote chars inside quotes are duplicated
        # escapechar = "\\"  # unused
        lineterminator = linesep
        quotechar = quote
        quoting = csv.QUOTE_MINIMAL
        skipinitialspace = False  # don't ignore spaces before a delimiter
        strict = False  # be permissive on malformed csv input

    return dialect
